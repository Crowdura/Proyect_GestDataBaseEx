import openpyxl
from openpyxl.styles import PatternFill 

from ...Funct_Esp.classFunct import classFunRecArc


class classRead:
    def __init__(self, dir, nam):
        id_user = 0
        lv_data = []
        self.dir = dir
        self.nam = nam
        self.lo_funcEsp = classFunRecArc(dir)
    def readText(self):
        lv_url = self.lo_funcEsp.reachArchiUni 
        print(lv_url)
        return
    def readExcel(self):

        lv_ubiArch = self.lo_funcEsp.reachArchiUni(self.nam)
        
        lv_worksheet = openpyxl.load_workbook(lv_ubiArch)

        
        lv_sheet = lv_worksheet['Sheet1']

        lv_cell  = lv_sheet['A2']
        
        lo_refColor = PatternFill(start_color="e54d2c", end_color="e54d2c", fill_type="solid")

        lv_cell.fill = lo_refColor

        lv_worksheet.save(lv_ubiArch)


        print(lv_sheet.calculate_dimension)
        print(lv_sheet['A2'].value)
        print(lv_sheet['A2'].fill.start_color.index)

        print(lv_ubiArch)
        return
"""lv_data = lv_workSheet['Sheet1']

        print(lv_data(lv_data['A1'].value))"""
        
